import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

def format_date_for_display(iso_date):
    """Преобразует ISO дату в ДД.ММ.ГГГГ."""
    if iso_date and re.match(r'^\d{4}-\d{2}-\d{2}$', iso_date):
        year, month, day = iso_date.split('-')
        return f"{day}.{month}.{year}"
    return iso_date

def create_excel_report(records):
    """Создаёт Excel-файл с заливкой колонок и границами ячеек."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Journal"

    # Заголовки
    headers = ['Дата-время приема', 'Препарат', 'Потенция', 'Дата события', 'Время события', 'Описание события']
    ws.append(headers)

    # Цвета
    green_fill = PatternFill(start_color="E6F0E6", end_color="E6F0E6", fill_type="solid")   # пастельно-зелёный
    blue_fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")    # пастельно-голубой
    coral_fill = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")   # светло-коралловый

    # Стили текста
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    event_alignment = Alignment(wrap_text=True, vertical='top')

    # Тонкие границы для всех ячеек
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Оформляем заголовки
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        # Первые три колонки – зелёные, остальные – голубые
        if col_idx <= 3:
            cell.fill = green_fill
        else:
            cell.fill = blue_fill

    # Заполняем данные
    for record in records:
        date_time = record.get('date-time', '')
        remedy = record.get('remedy', '')
        potency = record.get('potency', '')
        events = record.get('events', [])

        if events:
            for ev in events:
                event_date = format_date_for_display(ev.get('date', '')) if ev.get('date') else ''
                event_time = ev.get('time', '')
                description = ev.get('description', '')
                row = [date_time, remedy, potency, event_date, event_time, description]
                ws.append(row)
        else:
            row = [date_time, remedy, potency, '', '', '']
            ws.append(row)

    # Заливаем колонки дополнительных данных (4,5,6) коралловым и ставим границы
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=6):
        for cell in row:
            cell.fill = coral_fill
            cell.alignment = event_alignment  # перенос текста для описания
            cell.border = thin_border

    # Для основных колонок (1,2,3) добавляем только границы (без заливки)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.border = thin_border
            # Можно добавить заливку, если нужно, но оставим без

    # Автоподбор ширины колонок
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                try:
                    cell_value = str(cell.value)
                    lines = cell_value.split('\n')
                    for line in lines:
                        if len(line) > max_length:
                            max_length = len(line)
                except:
                    pass
        adjusted_width = max(max_length + 2, 10)
        ws.column_dimensions[col_letter].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
